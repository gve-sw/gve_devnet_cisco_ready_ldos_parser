import warnings


import os.path
import pathlib

import pandas as pd
from openpyxl.styles import Font

warnings.filterwarnings("ignore")


class Parser:
    DIR = f"{str(pathlib.Path().resolve())}"

    def __init__(
        self,
        path,
        output,
        start_date,
        end_date,
        include_minor_items,
        selected_date_target,
        file_type="single customer",
    ):
        self.customer_data = pd.read_excel(
            path,
            skiprows=3,
        )
        self._start_date = start_date
        self._end_date = end_date
        self.include_minor_items = include_minor_items
        self.selected_date_target = selected_date_target
        self.file_type = file_type

        self.result = []
        self.output = f"{self.DIR}/{output}"

        self._respect_date_interval()
        self._select_rel_columns()

        self.parse()

        self._setup_excel()
        self.write_to_excel("LDoS")
        self.write_to_excel("EoSMD")
        self.write_to_excel("EoPSD")
        self.write_to_excel("LRD")

    def _respect_date_interval(self):
        date_time_cols = [
            "Last Date of Support",
            "End of Product Sale Date",
            "End of Software Maintenance Date",
            "Last Renewal Date",
        ]
        for name in date_time_cols:
            self.customer_data[name] = pd.to_datetime(
                pd.to_numeric(self.customer_data[name]),
                origin="1899-12-30",
                unit="D",
            )
        # probably add some try except
        lower = self.customer_data[self.selected_date_target] > self._start_date
        upper = self.customer_data[self.selected_date_target] < self._end_date
        self.customer_data = self.customer_data[lower & upper]

    def parse(self):
        portfolio_lst = [
            "Security",
            "Collaboration Infrastructure",
            "Collaboration Endpoints",
            "Enterprise Routing",
            "Enterprise Switching",
            "Compute",
            "Wireless",
        ]
        self.result.append(self.build_for_management())
        for portfolio in portfolio_lst:
            self.result.append(self.build_portfolio(portfolio))

    def _select_rel_columns(self):
        rel_cols = [
            "Item Quantity",
            "Coverage",
            "Product ID",
            "Product Description",
            "End of Product Sale Date",
            "Last Renewal Date",
            "End of Software Maintenance Date",
            "Last Date of Support",
            "Business Entity",
            "Sub Business Entity",
            "Product Type",
            "Major/Minor",
            "Install Site GU Name",
        ]
        self.customer_data = self.customer_data[rel_cols]
        if self.include_minor_items == "no":
            self.customer_data = self.customer_data[
                self.customer_data["Major/Minor"] == "Major"
            ]

    def _grab_portfolio(self, portfolio):
        if portfolio == "Collaboration Endpoints":
            portfolio_df = self.customer_data.loc[
                self.customer_data["Sub Business Entity"].isin(
                    ["TP Endpoints", "UC Endpoints"]
                )
            ]
        elif portfolio == "Collaboration Infrastructure":
            portfolio_df = self.customer_data.loc[
                self.customer_data["Business Entity"] == "Collaboration"
            ]
            portfolio_df = portfolio_df.loc[
                ~portfolio_df["Sub Business Entity"].isin(
                    ["TP Endpoints", "UC Endpoints"]
                )
            ]
        elif portfolio == "Compute":
            portfolio_df = self.customer_data.loc[
                self.customer_data["Sub Business Entity"].isin(
                    ["Servers", "Hyper Converged"]
                )
            ]
        else:
            portfolio_df = self.customer_data.loc[
                self.customer_data["Business Entity"] == portfolio
            ]
        return portfolio_df

    def build_portfolio(self, portfolio):
        df_filter = ["Product ID", self.selected_date_target]  # "Last Date of Support"]
        portfolio_df = self._grab_portfolio(portfolio)
        agg_map = {
            "Item Quantity": "sum",
            "Coverage": "first",
            "Product ID": "first",
            "Product Description": "first",
            "Last Date of Support": "first",
            "End of Product Sale Date": "first",
            "End of Software Maintenance Date": "first",
            "Last Renewal Date": "first",
            "Business Entity": "first",
            "Sub Business Entity": "first",
            "Product Type": "first",
            "Major/Minor": "first",
            "Install Site GU Name": "first",
        }

        count_duplicates = (
            portfolio_df.groupby(df_filter).agg(agg_map).reset_index(drop=True)
        )
        count_duplicates.drop(
            columns=["Business Entity", "Sub Business Entity", "Product Type"],
            inplace=True,
        )
        if self.file_type == "single customer":
            count_duplicates.drop(columns=["Install Site GU Name"], inplace=True)

        count_duplicates.sort_values(by=self.selected_date_target, inplace=True)
        return (
            count_duplicates[count_duplicates[self.selected_date_target].notnull()],
            portfolio,
        )

    def build_for_management(self):
        columns = [
            "Item Quantity",
            "Coverage",
            "Product ID",
            "Product Description",
            "Last Date of Support",
            "End of Product Sale Date",
            "End of Software Maintenance Date",
            "Last Renewal Date",
            "Major/Minor",
            "Install Site GU Name",
        ]
        empty_df = pd.DataFrame(columns=columns)

        if self.file_type == "single customer":
            empty_df.drop(columns=["Install Site GU Name"], inplace=True)

        # return (pd.DataFrame(columns=columns), "Management Software")
        return (empty_df, "Management Software")

    def _setup_excel(self):
        if not os.path.isfile(self.output):
            writer = pd.ExcelWriter(self.output, engine="openpyxl")
            pd.DataFrame().to_excel(writer, sheet_name="LDoS", index=False)
            pd.DataFrame().to_excel(writer, sheet_name="EoSMD", index=False)
            pd.DataFrame().to_excel(writer, sheet_name="EoPSD", index=False)
            pd.DataFrame().to_excel(writer, sheet_name="LRD", index=False)
            writer.save()

    def write_to_excel(self, title):
        writer = pd.ExcelWriter(
            self.output, engine="openpyxl", mode="a", if_sheet_exists="overlay"
        )

        worksheet = writer.sheets[title]

        worksheet.column_dimensions["A"].width = 15
        worksheet.column_dimensions["B"].width = 12.5
        worksheet.column_dimensions["C"].width = 20
        worksheet.column_dimensions["D"].width = 59
        worksheet.column_dimensions["E"].width = 28
        worksheet.column_dimensions["F"].width = 15
        worksheet.column_dimensions["G"].width = 20
        cols_to_keep = [
            "Item Quantity",
            "Coverage",
            "Product ID",
            "Product Description",
            "End of Product Sale Date",
            "Last Renewal Date",
            "End of Software Maintenance Date",
            "Last Date of Support",
            "Major/Minor",
            "Install Site GU Name",
        ]
        if title == "LDoS":
            cols_to_keep.remove("Last Renewal Date")
            cols_to_keep.remove("End of Product Sale Date")
            cols_to_keep.remove("End of Software Maintenance Date")
            interest = "Last Date of Support"
        elif title == "LRD":
            cols_to_keep.remove("End of Product Sale Date")
            cols_to_keep.remove("End of Software Maintenance Date")
            cols_to_keep.remove("Last Date of Support")
            for df in self.result:
                df[0].sort_values(by="Last Renewal Date", inplace=True)
            interest = "Last Renewal Date"

        elif title == "EoPSD":
            cols_to_keep.remove("Last Renewal Date")
            cols_to_keep.remove("Last Date of Support")
            cols_to_keep.remove("End of Software Maintenance Date")
            for df in self.result:
                df[0].sort_values(by="End of Product Sale Date", inplace=True)
            interest = "End of Product Sale Date"
        elif title == "EoSMD":
            cols_to_keep.remove("Last Renewal Date")
            cols_to_keep.remove("End of Product Sale Date")
            cols_to_keep.remove("Last Date of Support")
            for df in self.result:
                df[0].sort_values(by="End of Software Maintenance Date", inplace=True)
            interest = "End of Software Maintenance Date"
        if self.file_type == "single customer":
            cols_to_keep.remove("Install Site GU Name")

        columns = [
            "Quantity",
            "Coverage",
            "Product ID",
            "Product Description",
            interest,
            "Major/Minor",
        ]

        if self.file_type == "multiple customers":
            columns.append("Install Site GU Name")

        col_idx = 1
        for col in columns:
            worksheet.cell(row=1, column=col_idx).value = col
            col_idx += 1

        row_num = 2

        bold_font = Font(bold=True)

        for df_category in self.result:
            df = df_category[0][cols_to_keep]
            try:
                # nice case when df is empty
                df[interest] = df[interest].dt.strftime("%d/%m/%Y")
            except:
                pass
            category = df_category[1]
            worksheet.cell(row=row_num, column=1).value = category
            worksheet.cell(row=row_num, column=1).font = bold_font
            df.to_excel(
                writer, sheet_name=title, startrow=row_num, index=False, header=None
            )
            row_num += df.shape[0] + 2

        writer.save()
